from flask import render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required
from app.laporan import laporan_bp
from app.laporan.forms import LaporanBarangForm, LaporanTransaksiForm, LaporanKontrakForm
from app.models import Barang, BarangMasuk, BarangKeluar, KontrakBarang, KategoriBarang, MerkBarang
from app.models.aset_tetap import AsetTetap
from app.models.merk_aset_tetap import MerkAsetTetap
from app.utils.excel_export import export_barang_to_excel, export_kontrak_to_excel, export_transaksi_to_excel
from app.utils.pdf_export import export_barang_to_pdf, export_kontrak_to_pdf, export_transaksi_to_pdf
from datetime import datetime


@laporan_bp.route('/')
@login_required
def index():
    """Halaman dashboard laporan"""
    return render_template('laporan/index.html', title='Laporan')


@laporan_bp.route('/barang')
@login_required
def laporan_barang():
    """Laporan daftar barang"""
    form = LaporanBarangForm()
    
    # Populate choices
    form.kategori_id.choices = [(0, 'Semua')] + [(k.id, k.nama_kategori) for k in KategoriBarang.query.all()]
    form.merk_id.choices = [(0, 'Semua')] + [(m.id, m.nama_merk) for m in MerkBarang.query.all()]
    
    # Query barang
    query = Barang.query
    
    # Apply filters
    if request.args.get('kategori_id') and int(request.args.get('kategori_id')) > 0:
        query = query.filter_by(kategori_id=int(request.args.get('kategori_id')))
        form.kategori_id.data = int(request.args.get('kategori_id'))
    
    if request.args.get('merk_id') and int(request.args.get('merk_id')) > 0:
        query = query.filter_by(merk_id=int(request.args.get('merk_id')))
        form.merk_id.data = int(request.args.get('merk_id'))
    
    barang_list = query.all()
    
    # Calculate stok akhir
    data = []
    for barang in barang_list:
        stok_akhir = barang.get_stok_akhir()
        
        # Filter by status
        if request.args.get('status'):
            status = request.args.get('status')
            if status == 'rendah' and stok_akhir >= 10:
                continue
            elif status == 'sedang' and (stok_akhir < 10 or stok_akhir >= 50):
                continue
            elif status == 'aman' and stok_akhir < 50:
                continue
        
        data.append({
            'barang': barang,
            'stok_akhir': stok_akhir
        })
    
    if request.args.get('status'):
        form.status.data = request.args.get('status')
    
    return render_template('laporan/barang.html', 
                          title='Laporan Barang',
                          form=form,
                          data=data)


@laporan_bp.route('/barang/export-excel')
@login_required
def export_barang_excel():
    """Export laporan barang ke Excel"""
    # Get filters from request
    query = Barang.query
    
    if request.args.get('kategori_id') and int(request.args.get('kategori_id')) > 0:
        query = query.filter_by(kategori_id=int(request.args.get('kategori_id')))
    
    if request.args.get('merk_id') and int(request.args.get('merk_id')) > 0:
        query = query.filter_by(merk_id=int(request.args.get('merk_id')))
    
    barang_list = query.all()
    
    # Calculate stok akhir
    data = []
    for barang in barang_list:
        stok_akhir = barang.get_stok_akhir()
        
        # Filter by status
        if request.args.get('status'):
            status = request.args.get('status')
            if status == 'rendah' and stok_akhir >= 10:
                continue
            elif status == 'sedang' and (stok_akhir < 10 or stok_akhir >= 50):
                continue
            elif status == 'aman' and stok_akhir < 50:
                continue
        
        data.append({
            'barang': barang,
            'stok_akhir': stok_akhir
        })
    
    # Generate Excel
    buffer = export_barang_to_excel(data)
    
    # Generate filename
    filename = f'Laporan_Barang_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(buffer, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/barang/export-pdf')
@login_required
def export_barang_pdf():
    """Export laporan barang ke PDF"""
    # Get filters from request
    query = Barang.query
    
    if request.args.get('kategori_id') and int(request.args.get('kategori_id')) > 0:
        query = query.filter_by(kategori_id=int(request.args.get('kategori_id')))
    
    if request.args.get('merk_id') and int(request.args.get('merk_id')) > 0:
        query = query.filter_by(merk_id=int(request.args.get('merk_id')))
    
    barang_list = query.all()
    
    # Calculate stok akhir
    data = []
    for barang in barang_list:
        stok_akhir = barang.get_stok_akhir()
        
        # Filter by status
        if request.args.get('status'):
            status = request.args.get('status')
            if status == 'rendah' and stok_akhir >= 10:
                continue
            elif status == 'sedang' and (stok_akhir < 10 or stok_akhir >= 50):
                continue
            elif status == 'aman' and stok_akhir < 50:
                continue
        
        data.append({
            'barang': barang,
            'stok_akhir': stok_akhir
        })
    
    # Generate PDF
    buffer = export_barang_to_pdf(data)
    
    # Generate filename
    filename = f'Laporan_Barang_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return send_file(buffer, 
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/kontrak')
@login_required
def laporan_kontrak():
    """Laporan daftar kontrak"""
    form = LaporanKontrakForm()
    
    # Populate tahun choices
    current_year = datetime.now().year
    form.tahun.choices = [(0, 'Semua')] + [(y, str(y)) for y in range(current_year - 5, current_year + 2)]
    
    # Query kontrak
    query = KontrakBarang.query
    
    # Apply filters
    if request.args.get('tahun') and int(request.args.get('tahun')) > 0:
        tahun = int(request.args.get('tahun'))
        query = query.filter(KontrakBarang.tanggal_kontrak.between(
            datetime(tahun, 1, 1),
            datetime(tahun, 12, 31)
        ))
        form.tahun.data = tahun
    
    if request.args.get('bulan'):
        bulan = int(request.args.get('bulan'))
        tahun = int(request.args.get('tahun', current_year))
        query = query.filter(KontrakBarang.tanggal_kontrak.between(
            datetime(tahun, bulan, 1),
            datetime(tahun, bulan, 28 if bulan == 2 else 30 if bulan in [4, 6, 9, 11] else 31)
        ))
        form.bulan.data = str(bulan)
    
    kontrak_list = query.order_by(KontrakBarang.tanggal_kontrak.desc()).all()
    
    return render_template('laporan/kontrak.html', 
                          title='Laporan Kontrak',
                          form=form,
                          kontrak_list=kontrak_list)


@laporan_bp.route('/kontrak/export-excel')
@login_required
def export_kontrak_excel():
    """Export laporan kontrak ke Excel"""
    # Get filters from request
    query = KontrakBarang.query
    
    if request.args.get('tahun') and int(request.args.get('tahun')) > 0:
        tahun = int(request.args.get('tahun'))
        query = query.filter(KontrakBarang.tanggal_kontrak.between(
            datetime(tahun, 1, 1),
            datetime(tahun, 12, 31)
        ))
    
    if request.args.get('bulan'):
        bulan = int(request.args.get('bulan'))
        tahun = int(request.args.get('tahun', datetime.now().year))
        query = query.filter(KontrakBarang.tanggal_kontrak.between(
            datetime(tahun, bulan, 1),
            datetime(tahun, bulan, 28 if bulan == 2 else 30 if bulan in [4, 6, 9, 11] else 31)
        ))
    
    kontrak_list = query.order_by(KontrakBarang.tanggal_kontrak.desc()).all()
    
    # Generate Excel
    buffer = export_kontrak_to_excel(kontrak_list)
    
    # Generate filename
    filename = f'Laporan_Kontrak_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(buffer, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/kontrak/export-pdf')
@login_required
def export_kontrak_pdf():
    """Export laporan kontrak ke PDF"""
    # Get filters from request
    query = KontrakBarang.query
    
    if request.args.get('tahun') and int(request.args.get('tahun')) > 0:
        tahun = int(request.args.get('tahun'))
        query = query.filter(KontrakBarang.tanggal_kontrak.between(
            datetime(tahun, 1, 1),
            datetime(tahun, 12, 31)
        ))
    
    if request.args.get('bulan'):
        bulan = int(request.args.get('bulan'))
        tahun = int(request.args.get('tahun', datetime.now().year))
        query = query.filter(KontrakBarang.tanggal_kontrak.between(
            datetime(tahun, bulan, 1),
            datetime(tahun, bulan, 28 if bulan == 2 else 30 if bulan in [4, 6, 9, 11] else 31)
        ))
    
    kontrak_list = query.order_by(KontrakBarang.tanggal_kontrak.desc()).all()
    
    # Generate PDF
    buffer = export_kontrak_to_pdf(kontrak_list)
    
    # Generate filename
    filename = f'Laporan_Kontrak_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return send_file(buffer, 
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/transaksi-masuk')
@login_required
def laporan_transaksi_masuk():
    """Laporan barang masuk"""
    form = LaporanTransaksiForm()
    
    # Query transaksi
    query = BarangMasuk.query
    
    # Apply filters
    if request.args.get('tanggal_awal'):
        tanggal_awal = datetime.strptime(request.args.get('tanggal_awal'), '%Y-%m-%d')
        query = query.filter(BarangMasuk.tanggal >= tanggal_awal)
        form.tanggal_awal.data = tanggal_awal
    
    if request.args.get('tanggal_akhir'):
        tanggal_akhir = datetime.strptime(request.args.get('tanggal_akhir'), '%Y-%m-%d')
        query = query.filter(BarangMasuk.tanggal <= tanggal_akhir)
        form.tanggal_akhir.data = tanggal_akhir
    
    if request.args.get('kode_barang'):
        query = query.filter(BarangMasuk.kode_barang.like(f"%{request.args.get('kode_barang')}%"))
        form.kode_barang.data = request.args.get('kode_barang')
    
    transaksi_list = query.order_by(BarangMasuk.tanggal.desc()).all()
    
    return render_template('laporan/transaksi_masuk.html', 
                          title='Laporan Barang Masuk',
                          form=form,
                          transaksi_list=transaksi_list)


@laporan_bp.route('/transaksi-masuk/export-excel')
@login_required
def export_transaksi_masuk_excel():
    """Export laporan barang masuk ke Excel"""
    # Get filters from request
    query = BarangMasuk.query
    
    tanggal_awal = None
    tanggal_akhir = None
    
    if request.args.get('tanggal_awal'):
        tanggal_awal = datetime.strptime(request.args.get('tanggal_awal'), '%Y-%m-%d')
        query = query.filter(BarangMasuk.tanggal >= tanggal_awal)
    
    if request.args.get('tanggal_akhir'):
        tanggal_akhir = datetime.strptime(request.args.get('tanggal_akhir'), '%Y-%m-%d')
        query = query.filter(BarangMasuk.tanggal <= tanggal_akhir)
    
    if request.args.get('kode_barang'):
        query = query.filter(BarangMasuk.kode_barang.like(f"%{request.args.get('kode_barang')}%"))
    
    transaksi_list = query.order_by(BarangMasuk.tanggal.desc()).all()
    
    # Generate Excel
    buffer = export_transaksi_to_excel(transaksi_list, 'masuk', tanggal_awal, tanggal_akhir)
    
    # Generate filename
    filename = f'Laporan_Barang_Masuk_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(buffer, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/transaksi-masuk/export-pdf')
@login_required
def export_transaksi_masuk_pdf():
    """Export laporan barang masuk ke PDF"""
    # Get filters from request
    query = BarangMasuk.query
    
    tanggal_awal = None
    tanggal_akhir = None
    
    if request.args.get('tanggal_awal'):
        tanggal_awal = datetime.strptime(request.args.get('tanggal_awal'), '%Y-%m-%d')
        query = query.filter(BarangMasuk.tanggal >= tanggal_awal)
    
    if request.args.get('tanggal_akhir'):
        tanggal_akhir = datetime.strptime(request.args.get('tanggal_akhir'), '%Y-%m-%d')
        query = query.filter(BarangMasuk.tanggal <= tanggal_akhir)
    
    if request.args.get('kode_barang'):
        query = query.filter(BarangMasuk.kode_barang.like(f"%{request.args.get('kode_barang')}%"))
    
    transaksi_list = query.order_by(BarangMasuk.tanggal.desc()).all()
    
    # Generate PDF
    buffer = export_transaksi_to_pdf(transaksi_list, 'masuk', tanggal_awal, tanggal_akhir)
    
    # Generate filename
    filename = f'Laporan_Barang_Masuk_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return send_file(buffer, 
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/transaksi-keluar')
@login_required
def laporan_transaksi_keluar():
    """Laporan barang keluar"""
    form = LaporanTransaksiForm()
    
    # Query transaksi
    query = BarangKeluar.query
    
    # Apply filters
    if request.args.get('tanggal_awal'):
        tanggal_awal = datetime.strptime(request.args.get('tanggal_awal'), '%Y-%m-%d')
        query = query.filter(BarangKeluar.tanggal >= tanggal_awal)
        form.tanggal_awal.data = tanggal_awal
    
    if request.args.get('tanggal_akhir'):
        tanggal_akhir = datetime.strptime(request.args.get('tanggal_akhir'), '%Y-%m-%d')
        query = query.filter(BarangKeluar.tanggal <= tanggal_akhir)
        form.tanggal_akhir.data = tanggal_akhir
    
    if request.args.get('kode_barang'):
        query = query.filter(BarangKeluar.kode_barang.like(f"%{request.args.get('kode_barang')}%"))
        form.kode_barang.data = request.args.get('kode_barang')
    
    transaksi_list = query.order_by(BarangKeluar.tanggal.desc()).all()
    
    return render_template('laporan/transaksi_keluar.html', 
                          title='Laporan Barang Keluar',
                          form=form,
                          transaksi_list=transaksi_list)


@laporan_bp.route('/transaksi-keluar/export-excel')
@login_required
def export_transaksi_keluar_excel():
    """Export laporan barang keluar ke Excel"""
    # Get filters from request
    query = BarangKeluar.query
    
    tanggal_awal = None
    tanggal_akhir = None
    
    if request.args.get('tanggal_awal'):
        tanggal_awal = datetime.strptime(request.args.get('tanggal_awal'), '%Y-%m-%d')
        query = query.filter(BarangKeluar.tanggal >= tanggal_awal)
    
    if request.args.get('tanggal_akhir'):
        tanggal_akhir = datetime.strptime(request.args.get('tanggal_akhir'), '%Y-%m-%d')
        query = query.filter(BarangKeluar.tanggal <= tanggal_akhir)
    
    if request.args.get('kode_barang'):
        query = query.filter(BarangKeluar.kode_barang.like(f"%{request.args.get('kode_barang')}%"))
    
    transaksi_list = query.order_by(BarangKeluar.tanggal.desc()).all()
    
    # Generate Excel
    buffer = export_transaksi_to_excel(transaksi_list, 'keluar', tanggal_awal, tanggal_akhir)
    
    # Generate filename
    filename = f'Laporan_Barang_Keluar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(buffer, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/transaksi-keluar/export-pdf')
@login_required
def export_transaksi_keluar_pdf():
    """Export laporan barang keluar ke PDF"""
    # Get filters from request
    query = BarangKeluar.query
    
    tanggal_awal = None
    tanggal_akhir = None
    
    if request.args.get('tanggal_awal'):
        tanggal_awal = datetime.strptime(request.args.get('tanggal_awal'), '%Y-%m-%d')
        query = query.filter(BarangKeluar.tanggal >= tanggal_awal)
    
    if request.args.get('tanggal_akhir'):
        tanggal_akhir = datetime.strptime(request.args.get('tanggal_akhir'), '%Y-%m-%d')
        query = query.filter(BarangKeluar.tanggal <= tanggal_akhir)
    
    if request.args.get('kode_barang'):
        query = query.filter(BarangKeluar.kode_barang.like(f"%{request.args.get('kode_barang')}%"))
    
    transaksi_list = query.order_by(BarangKeluar.tanggal.desc()).all()
    
    # Generate PDF
    buffer = export_transaksi_to_pdf(transaksi_list, 'keluar', tanggal_awal, tanggal_akhir)
    
    # Generate filename
    filename = f'Laporan_Barang_Keluar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return send_file(buffer, 
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/aset-tetap')
@login_required
def laporan_aset_tetap():
    """Laporan daftar aset tetap"""
    from app.laporan.forms import LaporanAsetTetapForm
    form = LaporanAsetTetapForm()
    
    # Populate choices
    form.kategori_id.choices = [(0, 'Semua')] + [(k.id, k.nama_kategori) for k in KategoriBarang.query.all()]
    form.merk_aset_tetap_id.choices = [(0, 'Semua')] + [(m.id, m.nama_merk) for m in MerkAsetTetap.query.all()]
    
    # Query aset tetap
    query = AsetTetap.query
    
    # Apply filters
    if request.args.get('kategori_id') and int(request.args.get('kategori_id')) > 0:
        query = query.filter_by(kategori_id=int(request.args.get('kategori_id')))
        form.kategori_id.data = int(request.args.get('kategori_id'))
    
    if request.args.get('merk_aset_tetap_id') and int(request.args.get('merk_aset_tetap_id')) > 0:
        query = query.filter_by(merk_aset_tetap_id=int(request.args.get('merk_aset_tetap_id')))
        form.merk_aset_tetap_id.data = int(request.args.get('merk_aset_tetap_id'))
    
    aset_list = query.order_by(AsetTetap.kode_aset).all()
    
    return render_template('laporan/aset_tetap.html', 
                          title='Laporan Aset Tetap',
                          form=form,
                          aset_list=aset_list)


@laporan_bp.route('/aset-tetap/export-excel')
@login_required
def export_aset_tetap_excel():
    """Export laporan aset tetap ke Excel"""
    # Get filters from request
    query = AsetTetap.query
    
    if request.args.get('kategori_id') and int(request.args.get('kategori_id')) > 0:
        query = query.filter_by(kategori_id=int(request.args.get('kategori_id')))
    
    if request.args.get('merk_aset_tetap_id') and int(request.args.get('merk_aset_tetap_id')) > 0:
        query = query.filter_by(merk_aset_tetap_id=int(request.args.get('merk_aset_tetap_id')))
    
    aset_list = query.order_by(AsetTetap.kode_aset).all()
    
    # Generate Excel using openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Aset Tetap'
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 20
    
    # Add header
    headers = ['No', 'Kode Aset', 'Nama Aset', 'Kategori', 'Merk', 'Kontrak/SPK', 'Tempat Penggunaan', 'Nama Pengguna']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add data
    for row, aset in enumerate(aset_list, 2):
        ws.cell(row=row, column=1).value = row - 1
        ws.cell(row=row, column=2).value = aset.kode_aset
        ws.cell(row=row, column=3).value = aset.nama_aset
        ws.cell(row=row, column=4).value = aset.kategori.nama_kategori if aset.kategori else '-'
        ws.cell(row=row, column=5).value = aset.merk_aset_tetap.nama_merk if aset.merk_aset_tetap else '-'
        ws.cell(row=row, column=6).value = aset.kontrak_spk or '-'
        ws.cell(row=row, column=7).value = aset.tempat_penggunaan or '-'
        ws.cell(row=row, column=8).value = aset.nama_pengguna or '-'
        
        # Apply borders and alignment
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename
    filename = f'Laporan_Aset_Tetap_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(buffer, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename)


@laporan_bp.route('/aset-tetap/export-pdf')
@login_required
def export_aset_tetap_pdf():
    """Export laporan aset tetap ke PDF"""
    # Get filters from request
    query = AsetTetap.query
    
    if request.args.get('kategori_id') and int(request.args.get('kategori_id')) > 0:
        query = query.filter_by(kategori_id=int(request.args.get('kategori_id')))
    
    if request.args.get('merk_aset_tetap_id') and int(request.args.get('merk_aset_tetap_id')) > 0:
        query = query.filter_by(merk_aset_tetap_id=int(request.args.get('merk_aset_tetap_id')))
    
    aset_list = query.order_by(AsetTetap.kode_aset).all()
    
    # Generate PDF using reportlab
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph('Laporan Daftar Aset Tetap', title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Prepare table data
    table_data = [['No', 'Kode Aset', 'Nama Aset', 'Kategori', 'Merk', 'Kontrak/SPK', 'Tempat Penggunaan', 'Nama Pengguna']]
    
    for idx, aset in enumerate(aset_list, 1):
        table_data.append([
            str(idx),
            aset.kode_aset,
            aset.nama_aset,
            aset.kategori.nama_kategori if aset.kategori else '-',
            aset.merk_aset_tetap.nama_merk if aset.merk_aset_tetap else '-',
            aset.kontrak_spk or '-',
            aset.tempat_penggunaan or '-',
            aset.nama_pengguna or '-'
        ])
    
    # Create table
    table = Table(table_data, colWidths=[0.4*inch, 1.0*inch, 1.5*inch, 1.0*inch, 1.0*inch, 1.2*inch, 1.5*inch, 1.2*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('WRAP', (0, 0), (-1, -1), True),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    # Generate filename
    filename = f'Laporan_Aset_Tetap_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return send_file(buffer, 
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=filename)
