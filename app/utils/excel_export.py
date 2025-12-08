"""Excel Export Utility using openpyxl"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO


class ExcelExporter:
    """Class untuk generate Excel reports"""
    
    def __init__(self, title="Laporan"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.title = title
        self.current_row = 1
        
        # Styles
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.bold_font = Font(name='Arial', size=10, bold=True)
        
        # Border
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def add_title(self, title, subtitle=None):
        """Tambah judul laporan"""
        self.ws.merge_cells(f'A{self.current_row}:F{self.current_row}')
        cell = self.ws[f'A{self.current_row}']
        cell.value = title
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.current_row += 1
        
        if subtitle:
            self.ws.merge_cells(f'A{self.current_row}:F{self.current_row}')
            cell = self.ws[f'A{self.current_row}']
            cell.value = subtitle
            cell.font = self.normal_font
            cell.alignment = Alignment(horizontal='center')
            self.current_row += 1
        
        self.current_row += 1  # Empty row
    
    def add_info(self, label, value):
        """Tambah info row (e.g., Periode: ...)"""
        self.ws[f'A{self.current_row}'] = label
        self.ws[f'A{self.current_row}'].font = self.bold_font
        self.ws[f'B{self.current_row}'] = value
        self.ws[f'B{self.current_row}'].font = self.normal_font
        self.current_row += 1
    
    def add_table_header(self, headers):
        """Tambah header tabel"""
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        self.current_row += 1
    
    def add_table_row(self, data, is_bold=False):
        """Tambah data row"""
        for col_num, value in enumerate(data, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num)
            cell.value = value
            cell.font = self.bold_font if is_bold else self.normal_font
            cell.border = self.border
            
            # Alignment untuk angka
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
        
        self.current_row += 1
    
    def add_summary_row(self, data):
        """Tambah row summary/total"""
        self.add_table_row(data, is_bold=True)
    
    def auto_adjust_columns(self):
        """Auto adjust column width"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_footer(self, text):
        """Tambah footer"""
        self.current_row += 1
        self.ws.merge_cells(f'A{self.current_row}:F{self.current_row}')
        cell = self.ws[f'A{self.current_row}']
        cell.value = text
        cell.font = Font(name='Arial', size=9, italic=True)
        cell.alignment = Alignment(horizontal='center')
    
    def save(self, filename):
        """Save to file"""
        self.auto_adjust_columns()
        self.wb.save(filename)
    
    def get_bytes(self):
        """Return as BytesIO for download"""
        self.auto_adjust_columns()
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output


# Helper functions untuk laporan spesifik
def export_barang_to_excel(barang_list, filename=None):
    """Export daftar barang ke Excel"""
    exporter = ExcelExporter()
    
    # Title
    exporter.add_title(
        "LAPORAN DAFTAR BARANG",
        f"Per {datetime.now().strftime('%d %B %Y')}"
    )
    
    # Info
    exporter.add_info("Total Barang:", f"{len(barang_list)} item")
    exporter.add_info("Dicetak pada:", datetime.now().strftime('%d/%m/%Y %H:%M'))
    exporter.current_row += 1
    
    # Table Header
    headers = ['No', 'Kode Barang', 'Nama Barang', 'Kategori', 'Merk', 'Satuan', 'Stok Awal', 'Stok Akhir', 'Status']
    exporter.add_table_header(headers)
    
    # Data Rows
    for idx, item in enumerate(barang_list, 1):
        barang = item['barang']
        stok_akhir = item['stok_akhir']
        
        # Status stok
        if stok_akhir < 10:
            status = "RENDAH"
        elif stok_akhir < 50:
            status = "SEDANG"
        else:
            status = "AMAN"
        
        row_data = [
            idx,
            barang.kode_barang,
            barang.nama_barang,
            barang.kategori.nama_kategori if barang.kategori else '-',
            barang.merk.nama_merk if barang.merk else '-',
            barang.satuan,
            barang.stok_awal,
            stok_akhir,
            status
        ]
        exporter.add_table_row(row_data)
    
    # Footer
    exporter.add_footer(f"Dibuat oleh Inven-Go System pada {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    if filename:
        exporter.save(filename)
    
    return exporter.get_bytes()


def export_kontrak_to_excel(kontrak_list, filename=None):
    """Export daftar kontrak ke Excel"""
    exporter = ExcelExporter()
    
    # Title
    exporter.add_title(
        "LAPORAN DAFTAR KONTRAK/SPK",
        f"Per {datetime.now().strftime('%d %B %Y')}"
    )
    
    # Info
    exporter.add_info("Total Kontrak:", f"{len(kontrak_list)} kontrak")
    exporter.add_info("Dicetak pada:", datetime.now().strftime('%d/%m/%Y %H:%M'))
    exporter.current_row += 1
    
    # Table Header
    headers = ['No', 'Nomor Kontrak', 'Tanggal', 'Jumlah Barang', 'Total Qty', 'Total Nilai (Rp)']
    exporter.add_table_header(headers)
    
    # Data Rows
    total_nilai = 0
    for idx, kontrak in enumerate(kontrak_list, 1):
        nilai_kontrak = kontrak.get_total_nilai()
        total_nilai += nilai_kontrak
        
        row_data = [
            idx,
            kontrak.nomor_kontrak,
            kontrak.tanggal_kontrak.strftime('%d/%m/%Y'),
            kontrak.barang_kontrak.count(),
            kontrak.get_total_qty(),
            f"{nilai_kontrak:,.0f}" if nilai_kontrak > 0 else '-'
        ]
        exporter.add_table_row(row_data)
    
    # Summary
    summary_data = ['', '', '', '', 'TOTAL:', f"{total_nilai:,.0f}"]
    exporter.add_summary_row(summary_data)
    
    # Footer
    exporter.add_footer(f"Dibuat oleh Inven-Go System pada {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    if filename:
        exporter.save(filename)
    
    return exporter.get_bytes()


def export_transaksi_to_excel(transaksi_list, jenis, tanggal_awal=None, tanggal_akhir=None, filename=None):
    """Export transaksi masuk/keluar ke Excel"""
    exporter = ExcelExporter()
    
    # Title
    title = f"LAPORAN BARANG {jenis.upper()}"
    if tanggal_awal and tanggal_akhir:
        subtitle = f"Periode: {tanggal_awal.strftime('%d/%m/%Y')} - {tanggal_akhir.strftime('%d/%m/%Y')}"
    else:
        subtitle = f"Per {datetime.now().strftime('%d %B %Y')}"
    
    exporter.add_title(title, subtitle)
    
    # Info
    exporter.add_info("Total Transaksi:", f"{len(transaksi_list)} transaksi")
    exporter.add_info("Dicetak pada:", datetime.now().strftime('%d/%m/%Y %H:%M'))
    exporter.current_row += 1
    
    # Table Header
    headers = ['No', 'Tanggal', 'Kode Barang', 'Nama Barang', 'Qty', 'Satuan', 'Keterangan']
    exporter.add_table_header(headers)
    
    # Data Rows
    total_qty = 0
    for idx, transaksi in enumerate(transaksi_list, 1):
        total_qty += transaksi.qty
        
        row_data = [
            idx,
            transaksi.tanggal.strftime('%d/%m/%Y'),
            transaksi.kode_barang,
            transaksi.barang.nama_barang if transaksi.barang else '-',
            transaksi.qty,
            transaksi.barang.satuan if transaksi.barang else '-',
            transaksi.keterangan or '-'
        ]
        exporter.add_table_row(row_data)
    
    # Summary
    summary_data = ['', '', '', '', total_qty, '', '']
    exporter.add_summary_row(summary_data)
    
    # Footer
    exporter.add_footer(f"Dibuat oleh Inven-Go System pada {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    if filename:
        exporter.save(filename)
    
    return exporter.get_bytes()
